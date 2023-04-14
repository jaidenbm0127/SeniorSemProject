import psutil
import GPUtil
import helpers as hlp
import constants
# from PerformanceMonitoring.process_data import ProcessData
from process_data import ProcessData
from openpyxl import Workbook


class Processor:
    # These dictionaries hold all the processes from the current run and the previous run, respectively.
    current_iteration_cpu = {}
    last_iteration_cpu = {}
    # Simply int to keep track of gpu_utilization
    gpu_utilization = None

    row = 1

    wb = Workbook()
    ws = wb.active

    # The functions that run all the other functions. Gets a dictionary of processes, checks GPU utilization, and
    # compares the current list of processes to the last runs.
    def process(self):

        Processor.get_processes(self)

        # Basically, if the last_iteration_cpu dictionary is empty,  that means this is the first run. So
        # just assign something to it and continue.
        if not bool(self.last_iteration_cpu):
            self.last_iteration_cpu = self.current_iteration_cpu
            return

        # This detects if the device has a GPU. If it does, it'll do GPU utilization checks.
        if len(GPUtil.getGPUs()) != 0:
            Processor.check_gpu_utilization(self)

        Processor.compare_current_iteration_to_last(self)
        Processor.compare_last_iteration_to_current(self)

        # At the  end of all comparisons, the current iteration of processes is now the last iteration
        # for the next run.
        self.last_iteration_cpu = self.current_iteration_cpu

    # Gets a dictionary of processes currently running on the device.
    def get_processes(self):

        # Make current_iteration_cpu an empty dictionary
        self.current_iteration_cpu = {}

        # Iterate over all running process
        for proc in psutil.process_iter():
            try:
                # Grab the cpu percent utilization and memory utilization of the current process
                cur_cpu_percent = round(proc.cpu_percent(interval=None) / psutil.cpu_count(), 2)
                cur_mem_percent = round(proc.memory_percent(), 2)
                if cur_cpu_percent > constants.CPU_THRESHOLD and cur_mem_percent > constants.MEMORY_THRESHOLD:
                    # Get process name, pid, cpu percent, memory percent, and priority.
                    process_name = proc.name()
                    process_id = proc.pid
                    process_cpu = cur_cpu_percent
                    process_memory = cur_mem_percent
                    process_score = proc.nice()

                    # Store into temp object to be added to dictionary of current iteration.
                    temp = ProcessData(process_name, process_cpu, process_memory, process_score)
                    self.current_iteration_cpu[process_id] = temp

            except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                pass

    def compare_last_iteration_to_current(self):
        for key in self.last_iteration_cpu:
            # If the key is in the last run and not in current_iteration_dict, we know that the process
            # has dropped off.
            if key not in self.current_iteration_cpu:
                # Keep pycharm from popping up a notification every time a new worker is made for the program.
                print(constants.PROCESS_REMOVED_MSG.format(self.last_iteration_cpu[key].process_id,
                                                           self.last_iteration_cpu[key].process_cpu,
                                                           self.last_iteration_cpu[key].process_memory,
                                                           self.last_iteration_cpu[key].process_priority))
            # If the key is in both dictionaries, that means that the process is in both dictionaries and thus,
            # still exists, and they need to be compared between runs.
            else:
                Processor.check_cpu_differences(self, key)
                Processor.check_memory_differences(self, key)

    def compare_current_iteration_to_last(self):
        for key in self.current_iteration_cpu:
            # If the key is in current_iteration_dict (current run) and not in the last run dict then that means the
            # process has been added
            if key not in self.last_iteration_cpu:
                print(constants.PROCESS_ADDED_MSG.format(self.current_iteration_cpu[key].process_id,
                                                         self.current_iteration_cpu[key].process_cpu,
                                                         self.current_iteration_cpu[key].process_memory,
                                                         self.current_iteration_cpu[key].process_priority))

    # Takes a key and compares the cpu utilization differences for that processes between last run and this run.
    def check_cpu_differences(self, key):

        last_iter_cpu = self.last_iteration_cpu[key].process_cpu
        current_iter_cpu = self.current_iteration_cpu[key].process_cpu

        # If the percentage change is greater than what CPU_PERCENTAGE_CHANGE_THRESHOLD is in constants.py,
        # print out a message.
        if hlp.calculate_percentage_difference(last_iter_cpu,
                                               current_iter_cpu) > constants.CPU_PERCENTAGE_CHANGE_THRESHOLD:
            self.ws['A1'] = 'CPU Utilization'
            self.ws['A' + str(self.row)] = hlp.calculate_percentage_difference(last_iter_cpu, current_iter_cpu)

            self.ws['B1'] = 'CPU Usage'
            self.ws['B' + str(self.row)] = current_iter_cpu

            # If last > current, that means CPU usage decreased.
            if last_iter_cpu > current_iter_cpu:
                print(constants.CPU_DECREASED_UTILIZATION_MSG.format(key, self.last_iteration_cpu[key].process_id,
                                                                     hlp.calculate_percentage_difference
                                                                     (current_iter_cpu, last_iter_cpu),
                                                                     round(current_iter_cpu, 2)
                                                                     ))
            # If last <>> current, that means CPU usage increase.
            elif last_iter_cpu < current_iter_cpu:
                print(constants.CPU_INCREASED_UTILIZATION_MSG.format(key, self.last_iteration_cpu[key].process_id,
                                                                     hlp.calculate_percentage_difference
                                                                     (current_iter_cpu, last_iter_cpu),
                                                                     round(current_iter_cpu, 2)
                                                                     ))

            self.row += 1

    # Takes a key and compares the memory utilization differences for that processes between last run and this run.
    # Functionally identical to check_cpu_differences
    def check_memory_differences(self, key):

        last_iter_mem = self.last_iteration_cpu[key].process_memory
        current_iter_mem = self.current_iteration_cpu[key].process_memory

        # If the percentage change is greater than what MEMORY_PERCENTAGE_CHANGE_THRESHOLD is in constants.py,
        # print out a message.
        if hlp.calculate_percentage_difference(last_iter_mem,
                                               current_iter_mem) > constants.MEMORY_PERCENTAGE_CHANGE_THRESHOLD:
            self.ws['C1'] = ' MEMORY utilization'
            self.ws['C' + str(self.row)] = hlp.calculate_percentage_difference(last_iter_mem, current_iter_mem)

            self.ws['D1'] = 'MEMORY usage'
            self.ws['D' + str(self.row)] = current_iter_mem
            # If last > current, that means memory usage decreased.
            if last_iter_mem > current_iter_mem:
                print(constants.MEMORY_DECREASED_UTILIZATION_MSG.format(key, self.last_iteration_cpu[key].process_id,
                                                                        hlp.calculate_percentage_difference
                                                                        (current_iter_mem, last_iter_mem),
                                                                        round(current_iter_mem, 2)
                                                                        ))
            # If last < current, that means memory usage decreased.
            elif last_iter_mem < current_iter_mem:
                print(constants.MEMORY_INCREASED_UTILIZATION_MSG.format(key, self.last_iteration_cpu[key].process_id,
                                                                        hlp.calculate_percentage_difference
                                                                        (current_iter_mem, last_iter_mem),
                                                                        round(current_iter_mem, 2)
                                                                        ))
            self.row += 1

    def check_gpu_utilization(self):

        # GPUtil.getGPUs(0)[0] is assuming that the current device has a discrete GPU and not an APU.
        # It multiplies by 100 to get a percentage on the GPU load.
        current_gpu_utilization = GPUtil.getGPUs()[0].load * 100

        # If this is the first run, assign is the current_gpu_utilization and exit
        if self.gpu_utilization is None:
            self.gpu_utilization = current_gpu_utilization
            return
        else:
            # If the last run and current runs gpu utilization is not 0, and the percentage change between the two
            # runs is > than GPU_PERCENTAGE_CHANGE_THRESHOLD in constants.py, print out a message that it decreased/
            # increased.
            if current_gpu_utilization != 0 and self.gpu_utilization != 0 and \
                    hlp.calculate_percentage_difference(current_gpu_utilization, self.gpu_utilization) > \
                    constants.GPU_PERCENTAGE_CHANGE_THRESHOLD:
                self.ws['E1'] = 'GPU utilization'
                self.ws['E' + str(self.row)] = hlp.calculate_percentage_difference(current_gpu_utilization,
                                                                                   self.gpu_utilization)
                self.ws['F1'] = 'GPU current utilization'
                self.ws['F' + str(self.row)] = self.gpu_utilization
                if current_gpu_utilization > self.gpu_utilization:
                    print(constants.GPU_DECREASED_UTILIZATION_MSG.format
                          (hlp.calculate_percentage_difference
                           (self.gpu_utilization, current_gpu_utilization),
                           round(current_gpu_utilization, 2)
                           ))
                else:
                    print(constants.GPU_INCREASED_UTILIZATION_MSG.format(
                        hlp.calculate_percentage_difference(self.gpu_utilization, current_gpu_utilization),
                        round(current_gpu_utilization, 2)
                    ))

                self.row += 1
        self.gpu_utilization = current_gpu_utilization

    def save(self):

        self.wb.save('sample.xlsx')
